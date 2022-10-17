from django.shortcuts import render
import openpyxl
from Audit.models import AuditTriple, Dataset
from users.models import CustomUser
from django.core.exceptions import PermissionDenied

# Create your views here.

context = {}

def upload_data_view(request):
    # Check is user is staff. Only users with is_staff attribute can upload_data new data sets
    if not request.user.is_staff:
        raise PermissionDenied()

    if "GET" == request.method:
       return render(request,'upload_data.html',context)
    
    else:

        user_list = [usr for usr in CustomUser.objects.all() if  not usr.is_staff]

        excel_file = request.FILES["excel_file"]

        name = request.POST.get("datasetname")

        name = Dataset.objects.create(name=name) 

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        excel_data = list()

        # iterating over the rows and
        # getting value from each cell in row
        # excel_data is list of lists containing the triples
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data) 

        # Remove duplicates from excel_data
        excel_data_no_dupes = list()

        [excel_data_no_dupes.append(row) for row in excel_data if row not in excel_data_no_dupes]

        # Each group gets different data to audit.
        #row_num = worksheet.max_row  
        row_num = len(excel_data_no_dupes)
        group_dataset_size = row_num  // 6 
        n = group_dataset_size 

        triple_data = list()

        # using list comprehension divide the list into 6 lists
        group_data_list = [excel_data_no_dupes[i:i + n] for i in range(0, row_num, n)]
        
        # Group 1
        for data in group_data_list[0]:
            usr = user_list[0]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])
            
        for data in group_data_list[0]:
            usr = user_list[1]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        # Group 2
        for data in group_data_list[1]:
            usr = user_list[2]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        for data in group_data_list[1]:
            usr = user_list[3]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        # Group 3
        for data in group_data_list[2]:
            usr = user_list[4]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        for data in group_data_list[2]:
            usr = user_list[5]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        # Group 4
        for data in group_data_list[3]:
            usr = user_list[6]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        for data in group_data_list[3]:
            usr = user_list[7]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        # Group 5
        for data in group_data_list[4]:
            usr = user_list[8]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        for data in group_data_list[4]:
            usr = user_list[9]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        # Group 6
        for data in group_data_list[5]:
            usr = user_list[10]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])

        for data in group_data_list[5]:
            usr = user_list[11]
            entA = data[0].split('|')[0] 
            rel = data[1]
            entB = data[2].split('|')[0] 
            triple_data.append([usr,entA,rel,entB])


        triple_object_list = list()
        for data in triple_data:
            triple_object_list.append(AuditTriple(
                dataset=name,
                user = data[0],
                entityA = data[1],
                relation = data[2],
                entityB = data[3],

            ))

        AuditTriple.objects.bulk_create(triple_object_list)
      
        return render(request, 'upload_data.html', context)
